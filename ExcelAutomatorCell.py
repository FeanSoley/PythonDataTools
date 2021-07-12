import openpyxl, os, xlsxwriter

#dir = r'C:\Users\Sean\Desktop\Elizabeth Script\Test Data'
#searchWord = 'example_value'
outputFile = 'Output Data.xlsx'
#sheetNumber = 0

def getFileList(dir):
    fileList = []
    files = os.listdir(dir)
    for file in files:
        if '.xls' in file or '.xlsx' in file:
            fileList.append(file)
    return fileList

def getDataFromFile(inputRow, inputCol, filename, sheetNumber):
    # Open workbook
    workbook = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    # open sheet based on user selected sheet
    sheet = workbook.worksheets[sheetNumber]
    # Get value
    returnValue = sheet.cell(inputRow, inputCol).value
    workbook.close()
    return returnValue

def printListToExcel(list, filename):
    # Open write workbook
    workbook = xlsxwriter.Workbook(filename)
    # Create new sheet
    worksheet = workbook.add_worksheet()
    
    # For each peice of data print file and data value
    currentRow = 0
    for file, data in list:
        worksheet.write(currentRow, 0, file)
        worksheet.write(currentRow, 1, data)
        currentRow += 1
    workbook.close()

if __name__ == '__main__':
    # Get inputs from user
    dir = input('Enter Folder Path: ')
    rowNumber = int(input('Enter row number (starting at 1): '))
    colNumber = int(input('Enter column number (A=1, B=2, etc): '))
    sheetNumber = int(input('Sheet Number (starting at 0): '))
    
    # Clean up folder path given from user, remove back slashes
    cleanedDir = dir.replace('\\', '/')

    # Get file list from folder path
    fileList = getFileList(dir)

    list = []
    # For each file get data based on given row and column
    for file in fileList:
        value = getDataFromFile(rowNumber, colNumber, dir+'/'+file, sheetNumber)
        list.append([file, value])
    
    # print data to excel file
    printListToExcel(list, outputFile)

    # Tell user file was created and wait for their input
    print('Created Excel File\n')

    end = input('Press Enter to continue')
