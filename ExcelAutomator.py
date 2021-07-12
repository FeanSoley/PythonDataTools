import openpyxl, os, xlsxwriter

#dir = r'C:\Users\Sean\Desktop\Elizabeth Script\Test Data'
#searchWord = 'example_value'
outputFile = 'Output Data.xlsx'
#sheetNumber = 0

dir = input('Enter Folder Path: ')
searchWord = input('Search Word: ')
sheetNumber = int(input('Sheet Number (starting at 0): '))
direction = input('Enter location of data from search word (l, r, u d) :')

def getFileList(dir):
    fileList = []
    files = os.listdir(dir)
    for file in files:
        if '.xls' in file or '.xlsx' in file:
            fileList.append(file)
    return fileList

def getDataFromFile(valueString, filename, sheetNumber, direction):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheet = workbook.worksheets[sheetNumber]
    row, col = findValueString(valueString, sheet)
    newRow = row
    newCol = col
    if direction == 'r':
        newCol = col + 1
    elif direction == 'l':
        newCol = col - 1
    elif direction == 'u':
        newRow = row - 1
    elif direction == 'd':
        newRow = row + 1
    else:
        print('WRONG DIRECTION')
    returnValue = sheet.cell(newRow, newCol).value
    workbook.close()
    return returnValue


def findValueString(valueString, sheet):
    numberRows = sheet.max_row
    numberCols = sheet.max_column
    for row in range(1, numberRows+1):
        for col in range(1, numberCols+1):
            currentCell = sheet.cell(row, col).value
            if currentCell == valueString:
                return row, col

def printListToExcel(list, filename):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()
    
    currentRow = 0
    for file, data in list:
        worksheet.write(currentRow, 0, file)
        worksheet.write(currentRow, 1, data)
        currentRow += 1
    
    workbook.close()

cleanedDir = dir.replace('\\', '/')

fileList = getFileList(dir)

list = []

for file in fileList:
    value = getDataFromFile(searchWord, dir+'/'+file, sheetNumber, direction)
    list.append([file, value])
    
printListToExcel(list, outputFile)

print('Created Excel File\n')

end = input('Press Enter to continue')
